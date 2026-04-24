import streamlit as st
import google.generativeai as genai
import json
import pandas as pd
import os
import concurrent.futures
import io
import openpyxl

st.set_page_config(page_title="Számla Mester Pro", page_icon="🧾", layout="wide")
st.title("Realign-Számlafeldolgozó (Javított verzió)")

# --- OLDALSÁV: BEÁLLÍTÁSOK ---
st.sidebar.header("⚙️ Beállítások")
API_KEY = st.sidebar.text_input("Gemini API kulcs:", type="password")

if not API_KEY:
    st.warning("Kérlek, add meg az API kulcsodat a bal oldali menüsávban!")
    st.stop()

genai.configure(api_key=API_KEY)

try:
    valid_models = [m.name.replace('models/', '') for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
    alap_index = valid_models.index('gemini-1.5-flash-latest') if 'gemini-1.5-flash-latest' in valid_models else 0
    selected_model_name = st.sidebar.selectbox("AI Modell:", valid_models, index=alap_index)
    model = genai.GenerativeModel(selected_model_name)
except:
    st.error("Hiba az API kulccsal!")
    st.stop()

# --- FŐOLDAL ---
master_file = st.file_uploader("1. Lépés: Mester táblázat feltöltése (.xlsx)", type=["xlsx"])
uploaded_files = st.file_uploader("2. Lépés: Új számlák feltöltése", type=["pdf", "jpg", "jpeg", "png"], accept_multiple_files=True)

def process_invoice(uploaded_file, prompt, model):
    filename = uploaded_file.name
    temp_path = f"temp_{filename}"
    try:
        with open(temp_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        sample_file = genai.upload_file(path=temp_path)
        response = model.generate_content([sample_file, prompt])
        clean_json = response.text.replace('```json', '').replace('```', '').strip()
        return json.loads(clean_json)
    except Exception as e:
        return {"Szállító": f"HIBA: {filename}"}
    finally:
        if os.path.exists(temp_path):
            try: os.remove(temp_path)
            except: pass

if uploaded_files and st.button("Feldolgozás és Biztonságos Mentés"):
    all_new_rows = []
    prompt = """
    Elemezd a számlát és adj JSON választ. 
    Keresd: Szállító, Számlaszám, Számla kelte, Teljesítés dátuma, Fizetési határidő, Kifizetés hónapja, Nettó, Áfa, Bruttó, Pénznem, Nettó HUF, Áfa HUF, EUR fx.
    Ha valami nincs meg, írd: "HIBA: Nem található".
    """
    
    with st.spinner("Számlák elemzése..."):
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            futures = [executor.submit(process_invoice, f, prompt, model) for f in uploaded_files]
            all_new_rows = [f.result() for f in concurrent.futures.as_completed(futures)]

    new_df = pd.DataFrame(all_new_rows)

    if master_file:
        # BIZTONSÁGOS APPEND LOGIKA
        wb = openpyxl.load_workbook(master_file, data_only=False) # data_only=False megtartja a képleteket
        ws = wb.active
        
        # Megkeressük a fejléceket az első sorban
        headers = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
        
        # Meghatározzuk az utolsó valódi sort (ahol az A oszlop nem üres)
        last_row = ws.max_row
        while last_row > 1 and ws.cell(row=last_row, column=1).value is None:
            last_row -= 1
        
        # Adatok beírása oszlopnév alapján
        for index, row_data in new_df.iterrows():
            current_row = last_row + index + 1
            for key, value in row_data.items():
                if key in headers:
                    ws.cell(row=current_row, column=headers[key]).value = value
        
        output = io.BytesIO()
        wb.save(output)
        final_data = output.getvalue()
        st.success("Adatok hozzáadva az oszlopnevek alapján!")
    else:
        output = io.BytesIO()
        new_df.to_excel(output, index=False)
        final_data = output.getvalue()

    st.dataframe(new_df)
    st.download_button("⬇️ Letöltés", data=final_data, file_name="javitott_konyveles.xlsx")
